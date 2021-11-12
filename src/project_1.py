import itertools
from functools import partial

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy.optimize import minimize, NonlinearConstraint, Bounds

from solver import Solver


def optimize_v1(x0, E, W, D, xi=None, plot=True, u=True):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi, uniformCrossSection=u)

    solver = Solver(nodeCords, elemNodes, modE, Area, DispCon, Fval)
    solver.solve()

    print(solver.Stress)

    if plot:
        solver.plot()

    return calculate_c(x0, E, W, D, xi, u=u)


def parameter_calculation(x0, E=1, W=0, D=0, xi=None, uniformCrossSection=True):
    if xi is None:
        if uniformCrossSection:
            node1_x, node2_x, node2_y, node3_y, cross_section_width = x0
        else:
            node1_x, node2_x, node2_y, node3_y, cross_section_width = x0[0], x0[1], x0[2], x0[3], x0[4:]

        area = np.power(cross_section_width, 2)
    else:
        node1_x, node2_x, node2_y, node3_y = x0

        area = xi ** 2

    nodeCords = np.array([[node1_x, 0.0],
                          [node2_x, node2_y],
                          [0, node3_y],

                          [-node1_x, 0.0],
                          [-node2_x, node2_y],
                          ])
    # Nodal Coordinates
    elemNodes = np.array([[0, 1], [0, 2], [0, 4], [1, 2], [1, 3], [1, 4], [2, 3], [2, 4],
                          [3, 4]])  # Element connectivity: near node and far node

    modE = np.full((elemNodes.shape[0], 1), E)  # Young's modulus

    if uniformCrossSection:
        Area = np.full((elemNodes.shape[0], 1), area)  # cross section area
    else:
        Area = np.reshape(area, (elemNodes.shape[0], 1))

    # Displacement constraints # node number, degree of freedom (x=1, y=2), displacement value
    DispCon = np.array([[0, 1, 0.0], [0, 2, 0.0], [3, 1, 0.0],
                        [3, 2,
                         0.0]])

    Fval = np.array([[2, 2, -W], [2, 1, D]])  # Applied force # node number, orientation (x=1, y=2), value

    return nodeCords, elemNodes, modE, Area, DispCon, Fval


def stress_calculator(x0, E, W=0, D=0, xi=None, u=True):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi, uniformCrossSection=u)

    solver = Solver(nodeCords, elemNodes, modE, Area, DispCon, Fval)
    result, _ = solver.solve()

    return result


def critical_buckling(x0, E, W=0, D=0, safety_factor=1, xi=None, u=True):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi, uniformCrossSection=u)

    solver = Solver(nodeCords, elemNodes, modE, Area, DispCon, Fval)
    solver.solve()
    internal = solver.internal_forces
    critical = solver.critical_loads
    results = critical / safety_factor - np.abs(internal)

    return results.ravel()


def node_distance(x0, E=1, W=0, D=0, xi=None, u=True):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi, uniformCrossSection=u)

    t = [i for i in range(len(nodeCords))]
    c = list(itertools.combinations(t, 2))

    lengths = []

    for i, f in c:
        l_ = np.linalg.norm(nodeCords[i] - nodeCords[f])
        lengths.append(l_)

    return lengths


def calculate_c(x0, E=1, W=0, D=0, xi=None, u=True):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi, uniformCrossSection=u)

    sum_ = 0

    for i in range(elemNodes.shape[0]):
        first, last = elemNodes[i]
        l_i = np.linalg.norm(nodeCords[first] - nodeCords[last])
        a_i = Area[i]

        sum_ += a_i * l_i

    return nodeCords.shape[0] * sum_


def test():
    W = 80 * 1000
    D = 78771
    E = 200 * 1e9
    MaxYield = 600 * 1e6

    x = [10, 8, 20, 50, 1]
    output = optimize_v1(x, E, W, D)
    print(output)


def slenderness_ratio(x0, E=1, W=0, D=0, xi=None, u=True):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi, uniformCrossSection=u)

    slenderness = []

    for i in range(elemNodes.shape[0]):
        first, last = elemNodes[i]
        l_i = np.linalg.norm(nodeCords[first] - nodeCords[last])
        a_i = Area[i]
        slenderness.append((l_i * np.sqrt(12 / a_i))[0])

    return slenderness


def scipy_optimization(xi=None, uniform=True):
    W = 80 * 1000
    D = 78771
    E = 200 * 1e9
    MaxYield = 600 * 1e6
    safety_factor = 4

    if uniform:
        x0 = [3.59164903, 4.13653055, 4.27378281, 50.00000018, 0.34730275]  # 136.68311266
        l_b = [2.5, 0, 0, 50, 0]
        u_b = [10, np.inf, np.inf, np.inf, np.inf]
    else:
        x0 = [3.59164903, 4.13653055, 4.27378281, 50.00000018, 0.34730275, 0.34730275, 0.34730275, 0.34730275,
              0.34730275, 0.34730275, 0.34730275, 0.34730275, 0.34730275]  # 136.68311266

        x0 = [5.65215739, 3.59013432, 17.07125767, 50.00494497, 0.19227025,
              0.36781726, 0.1420632, 0.23090128, 0.14160986, 0.0684492,
              0.34904794, 0.23019538, 0.15407471]  # 91.59849992

        x0 = [8.21173603e+00, 3.41736976e+00, 2.08411552e+01, 5.00000000e+01,
              1.48163110e-01, 3.51050945e-01, 1.65349065e-01, 2.03401080e-01,
              1.65349065e-01, 4.73524913e-02, 3.51050944e-01, 2.03401080e-01,
              1.48163106e-01]  # 85.8863339

        l_b = [2.5, 0, 0, 50, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        u_b = [10, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf]

    if xi is None:
        bounds = Bounds(l_b, u_b)
    else:
        x0 = x0[:-1]

        bounds = Bounds(l_b[:-1], u_b[:-1])

    nlc = NonlinearConstraint(partial(stress_calculator, E=E, W=W, D=D, xi=xi, u=uniform), -MaxYield / safety_factor,
                              MaxYield / safety_factor, jac='3-point')

    tol = 0

    nlc2 = NonlinearConstraint(partial(critical_buckling, E=E, W=W, D=D, safety_factor=safety_factor, xi=xi, u=uniform),
                               0 + tol,
                               np.inf,
                               jac='3-point')
    #
    nlc3 = NonlinearConstraint(partial(node_distance, E=E, W=W, D=D, xi=xi, u=uniform), 2, np.inf, jac='3-point')

    nlc4 = NonlinearConstraint(partial(slenderness_ratio, E=E, W=W, D=D, xi=xi, u=uniform), 0, 500 - tol, jac='3-point')

    print("------------ INITIAL ------------")
    print(f'{x0=}')
    print(optimize_v1(x0, E, W, D, xi=xi, plot=False, u=uniform))
    print(node_distance(x0, E, W, D, xi=xi, u=uniform))
    print(slenderness_ratio(x0, E, W, D, xi=xi, u=uniform))

    print("------------ FINAL ------------")
    res = minimize(partial(calculate_c, E=E, W=W, D=D, xi=xi, u=uniform), x0, method='trust-constr', bounds=bounds,
                   constraints=[nlc, nlc2, nlc3, nlc4],
                   options={'verbose': 1, 'maxiter': 1 * 1e5}, jac='3-point')

    with open('../out.txt', 'w') as f:
        f.write(str(res))

    print(f'{res.x=}')
    print(optimize_v1(res.x, E, W, D, xi=xi, u=uniform))
    print(node_distance(res.x, E, W, D, xi=xi, u=uniform))
    print(stress_calculator(res.x, E, W, D, xi=xi, u=uniform))
    print(slenderness_ratio(res.x, E, W, D, xi=xi, u=uniform))
    print(critical_buckling(res.x, E, W, D, xi=xi, u=uniform))

    save_to_excel(res.x, E, W, D, xi, uniform)


def save_to_excel(x0, E, W, D, xi=None, uniform=True, file="../InputFea_a.xlsx"):
    nodeCords, elemNodes, modE, Area, DispCon, Fval = parameter_calculation(x0, E, W, D, xi,
                                                                            uniformCrossSection=uniform)

    solver = Solver(nodeCords, elemNodes, modE, Area, DispCon, Fval)
    solver.solve()

    workbook: Workbook = load_workbook(filename=file)

    sheet: Worksheet = workbook['elemNodes']

    for i in range(elemNodes.shape[0]):
        sheet.cell(i + 2, 1).value = elemNodes[i][0]
        sheet.cell(i + 2, 2).value = elemNodes[i][1]

    sheet: Worksheet = workbook['nodeCords']
    for i in range(nodeCords.shape[0]):
        sheet.cell(i + 2, 1).value = i
        sheet.cell(i + 2, 2).value = nodeCords[i][0]
        sheet.cell(i + 2, 3).value = nodeCords[i][1]

    sheet: Worksheet = workbook['CrossSection']
    for i in range(Area.shape[0]):
        sheet.cell(i + 2, 1).value = Area[i][0]
        sheet.cell(i + 2, 2).value = modE[i][0]

    sheet: Worksheet = workbook['DispCon']
    for i in range(DispCon.shape[0]):
        sheet.cell(i + 2, 1).value = DispCon[i][0]
        sheet.cell(i + 2, 2).value = DispCon[i][1]
        sheet.cell(i + 2, 3).value = DispCon[i][2]

    sheet: Worksheet = workbook['forces']
    for i in range(Fval.shape[0]):
        sheet.cell(i + 2, 1).value = Fval[i][0]
        sheet.cell(i + 2, 2).value = Fval[i][1]
        sheet.cell(i + 2, 3).value = Fval[i][2]

    sheet: Worksheet = workbook['EStress']
    for i in range(solver.Stress.shape[0]):
        index = i + 2

        sheet.cell(index, 2).value = solver.Stress[i]
        sheet.cell(index,
                   3).value = f'=IF(ISNUMBER(elemNodes!A{index}),SQRT((VLOOKUP(elemNodes!A{index},nodeCords!$A$2:$C$105,2,FALSE)-VLOOKUP(elemNodes!B{index},nodeCords!$A$2:$C$105,2,FALSE))^2+(VLOOKUP(elemNodes!A{index},nodeCords!$A$2:$C$105,3,FALSE)-VLOOKUP(elemNodes!B{index},nodeCords!$A$2:$C$105,3,FALSE))^2),"")'
        sheet.cell(index, 4).value = f'=IF(ISNUMBER(B{index}),B{index}*CrossSection!A{index},"")'
        sheet.cell(index,
                   5).value = f'=IF(ISNUMBER(B{index}),-(PI()^2*CrossSection!B{index}*CrossSection!A{index}^2)/(48*C{index}^2),"")'
        sheet.cell(index, 6).value = f'=IF(ISNUMBER(B{index}),C{index}*SQRT(12/CrossSection!$A{index}),"")'
        sheet.cell(index, 7).value = f'=IF(ISNUMBER(B{index}),IF(ABS(D{index})>ABS(E{index}),"X","J"),"")'
        sheet.cell(index, 8).value = f'=IF(ISNUMBER(B{index}),IF(ABS(B{index})>150000000,"X","J"),"")'
        sheet.cell(index, 9).value = f'=IF(ISNUMBER(B{index}),IF(F{index}>500,"X","J"),"")'
        sheet.cell(index, 10).value = f'=IF(ISNUMBER(B{index}),C{index}*CrossSection!$A{index},"")'

    workbook.save(file)
    print("Solved")


if __name__ == '__main__':
    # xi = 2.30056943e-02
    xi = None

    scipy_optimization(xi=xi)
